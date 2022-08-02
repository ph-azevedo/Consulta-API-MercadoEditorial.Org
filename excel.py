from openpyxl import Workbook, load_workbook
from main import MEGet, isbn, arquivo

try:
    planilhaexiste = True
    wb = load_workbook(f'{arquivo}.xlsx')
except:
    if FileNotFoundError:
        planilhaexiste = False



def adiciona_planilha():
    if planilhaexiste == False:
        wb = Workbook()
        ws = wb['Sheet']
        ws.append(['ISBN', 'Título', 'Editora', 'Edição', 'Ano de Lançamento', 'Preço'])
    elif planilhaexiste == True:
        wb = load_workbook('Teste.xlsx')
        ws = wb['Sheet']
        for x in ws['A']:
            if x.value == isbn:
                print('Item já está na lista.')
                exit()

    ws.append([isbn, MEGet.get_titulo(), MEGet.get_editora(), MEGet.get_edicao(), MEGet.get_ano(), MEGet.get_preco()])
    wb.save(f'{arquivo}.xlsx')

adiciona_planilha()